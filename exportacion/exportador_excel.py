from __future__ import annotations
from typing import List, Dict, Tuple, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from ..core.models import ScheduleResult, Process, ExecSlice

class ExportadorExcel:
    # =============== Utilidades internas para tablas =================
    def _escribir_tabla(self, ws, filas: List[Any]):
        if not filas:
            return
        primera = filas[0]
        if isinstance(primera, dict):
            headers = list(primera.keys())
            ws.append(headers)
            for fila in filas:
                ws.append([fila.get(h, "") for h in headers])
        else:
            for fila in filas:
                ws.append(list(fila))

    def _auto_ajustar_ancho(self, ws):
        for col in ws.columns:
            try:
                letter = col[0].column_letter
            except Exception:
                continue
            maxlen = 0
            for cell in col:
                v = cell.value
                if v is None:
                    continue
                maxlen = max(maxlen, len(str(v)))
            ws.column_dimensions[letter].width = min(maxlen + 2, 60)

    # ================== Modo clásico: sólo tablas ==================
    def exportar(self, nombre_archivo: str, hojas: List[tuple[str, List[Any]]]):
        wb = Workbook()
        wb.remove(wb.active)
        for nombre, datos in hojas:
            ws = wb.create_sheet(title=str(nombre)[:31])
            self._escribir_tabla(ws, datos)
            self._auto_ajustar_ancho(ws)
        wb.save(nombre_archivo)

    # ================== Gantt con grilla ==================
    def exportar_con_gantt(
        self,
        nombre_archivo: str,
        resultado: ScheduleResult,
        hojas: Optional[List[tuple[str, List[Any]]]] = None,
        procesos: Optional[List[Process]] = None,
        opciones: Optional[dict] = None,
    ):
        """
        Crea primero la hoja 'Gantt (grilla)' (activa) y luego las tablas opcionales.
        """
        wb = Workbook()
        wb.remove(wb.active)

        # 1) Gantt primero y dejarlo activo
        ws_gantt = self._crear_hoja_gantt_grilla(wb, resultado, procesos, opciones or {})
        wb.active = wb.worksheets.index(ws_gantt)

        # 2) Hojas tabulares
        if hojas:
            for nombre, datos in hojas:
                ws = wb.create_sheet(title=str(nombre)[:31])
                self._escribir_tabla(ws, datos)
                self._auto_ajustar_ancho(ws)

        wb.save(nombre_archivo)

    # ------------------- Helpers de Gantt -------------------
    def _orden_por_aparicion(self, resultado: ScheduleResult):
        """
        Orden base: por primera aparición en el timeline.
        Este orden lo usa el canvas para dibujar *de abajo hacia arriba*.
        """
        orden = []
        for sl in resultado.timeline:
            base = sl.process.replace("_BLOCK", "")
            if base not in orden:
                orden.append(base)
        # fallback: por métricas
        for pid in list(resultado.turnaround.keys()):
            if pid not in orden:
                orden.append(pid)
        return orden

    def _paleta(self):
        # hex sin '#'
        return [
            "4472C4", "ED7D31", "70AD47", "FFC000", "5B9BD5",
            "A5A5A5", "264478", "9E480E", "636363", "997300",
            "255E91", "43682B",
        ]

    def _io_intervals_precisos(
        self, resultado: ScheduleResult, procesos: List[Process]
    ) -> Dict[str, List[Tuple[int, int]]]:
        """
        Reconstrucción de E/S desde timeline:
        - Un ExecSlice con process terminando en "_BLOCK" representa un bloqueo.
        """
        io: Dict[str, List[Tuple[int, int]]] = {}
        for sl in resultado.timeline:
            if sl.process.endswith("_BLOCK"):
                base = sl.process.replace("_BLOCK", "")
                io.setdefault(base, []).append((int(sl.start), int(sl.end)))
        return io

    def _crear_hoja_gantt_grilla(
        self, wb: Workbook, resultado: ScheduleResult, procesos: Optional[List[Process]], op: dict
    ):
        """
        Dibuja la grilla con el mismo orden visual que el canvas.
        """
        ws = wb.create_sheet(title="Gantt (grilla)")

        # Config
        tiempo_total = int(max((sl.end for sl in resultado.timeline), default=0))

        # === Orden igual al canvas ===
        orden_canvas = self._orden_por_aparicion(resultado)  # bottom-up en canvas
        orden = list(reversed(orden_canvas))                 # top-down en Excel

        paleta = self._paleta()
        color_cpu_por_pid: Dict[str, str] = op.get("color_cpu_por_pid", {})
        pattern_cpu = op.get("pattern_cpu", "solid")
        color_io = op.get("color_io", "000000")
        pattern_io = op.get("pattern_io", "solid")
        fuente_cpu = Font(bold=True, color="FFFFFF")
        fuente_io = Font(color="FFFFFF", size=8)

        # mapping de colores
        for i, pid in enumerate(orden_canvas):
            color_cpu_por_pid.setdefault(pid, paleta[i % len(paleta)])

        # Encabezado
        ws["A1"] = "Proceso"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(bold=True)

        primera_fila_barras = 2
        ultima_fila_barras = primera_fila_barras + len(orden) - 1
        fila_tiempo = max(primera_fila_barras, ultima_fila_barras) + 1

        # Estilos
        thin = Side(style="thin", color="000000")
        border_thin = Border(top=thin, bottom=thin, left=thin, right=thin)

        # Anchos/altos
        ws.column_dimensions["A"].width = 12
        max_digits = len(str(tiempo_total))
        col_width = max(3, max_digits + 1)
        for t in range(tiempo_total):
            ws.column_dimensions[get_column_letter(2 + t)].width = col_width

        # Segmentos CPU ordenados
        segs_por_pid: Dict[str, List[Tuple[int, int]]] = {}
        for sl in resultado.timeline:
            if not sl.process.endswith("_BLOCK"):
                segs_por_pid.setdefault(sl.process, []).append((int(sl.start), int(sl.end)))
        for pid in segs_por_pid:
            segs_por_pid[pid].sort()

        # E/S
        io_por_pid = self._io_intervals_precisos(resultado, procesos) if procesos else {}

        # Llenar grilla vacía con bordes
        for r in range(primera_fila_barras, fila_tiempo + 1):
            for c in range(2, 2 + tiempo_total):
                cell = ws.cell(row=r, column=c)
                cell.border = border_thin
                cell.alignment = Alignment(horizontal="center", vertical="center")


        # --- Eje de tiempo ---
        if tiempo_total >= 0:
            c0 = 2
            cell0 = ws.cell(row=fila_tiempo, column=c0, value=0)
            cell0.alignment = Alignment(horizontal="left", vertical="center")
            cell0.border = border_thin
        for t in range(1, tiempo_total + 1):
            col = 2 + (t - 1)   # última columna será 2 + tiempo_total - 1
            cell = ws.cell(row=fila_tiempo, column=col, value=t)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = border_thin

        # Borde derecho columna A
        for r in range(primera_fila_barras, fila_tiempo + 1):
            ca = ws.cell(row=r, column=1)
            ca.border = Border(right=thin)
            ca.alignment = Alignment(horizontal="center", vertical="center")

        def _merge_segments(segments):
            """Une segmentos contiguos o solapados en una sola tupla (ini, fin)."""
            if not segments:
                return []
            segments = sorted(segments, key=lambda x: x[0])
            merged = [segments[0]]
            for ini, fin in segments[1:]:
                last_ini, last_fin = merged[-1]
                if ini <= last_fin:  # contiguo o solapado
                    merged[-1] = (last_ini, max(last_fin, fin))
                else:
                    merged.append((ini, fin))
            return merged

        # Pintar barras
        for idx, pid in enumerate(orden):
            fila = primera_fila_barras + idx
            ws.cell(row=fila, column=1, value=pid).alignment = Alignment(horizontal="center", vertical="center")

            # CPU
            color_cpu = color_cpu_por_pid.get(pid, paleta[orden_canvas.index(pid) % len(paleta)])
            fill_cpu = PatternFill(start_color=color_cpu, end_color=color_cpu, fill_type=pattern_cpu)

            # Unir segmentos contiguos antes de pintarlos
            for (ini, fin) in _merge_segments(segs_por_pid.get(pid, [])):
                if fin <= ini:
                    continue
                c1 = 2 + ini
                c2 = 2 + fin - 1
                ws.merge_cells(start_row=fila, start_column=c1, end_row=fila, end_column=c2)
                cell = ws.cell(row=fila, column=c1, value=pid)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = fuente_cpu
                cell.fill = fill_cpu
                for c in range(c1, c2 + 1):
                    ws.cell(row=fila, column=c).border = border_thin

            
                        # Arrival (sombra gris antes de que llegue el proceso)
            if procesos:
                proc = next((p for p in procesos if p.name == pid), None)
                if proc and proc.arrival > 0:
                    fill_arrival = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
                    for t in range(proc.arrival):
                        col = 2 + t
                        cell = ws.cell(row=fila, column=col)
                        cell.fill = fill_arrival
                        cell.border = border_thin


            # I/O
            if pid in io_por_pid:
                fill_io = PatternFill(start_color=color_io, end_color=color_io, fill_type=pattern_io)
                for (ini, fin) in io_por_pid[pid]:
                    if fin <= ini:
                        continue
                    c1 = 2 + ini
                    c2 = 2 + fin - 1
                    ws.merge_cells(start_row=fila, start_column=c1, end_row=fila, end_column=c2)
                    cell = ws.cell(row=fila, column=c1, value="IO")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = fuente_io
                    cell.fill = fill_io
                    for c in range(c1, c2 + 1):
                        ws.cell(row=fila, column=c).border = border_thin

        # Alturas
        ws.row_dimensions[1].height = 18
        for r in range(primera_fila_barras, fila_tiempo + 1):
            ws.row_dimensions[r].height = 18

        return ws
